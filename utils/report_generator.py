import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils.logger import log_info, log_error

def format_excel_sheet(ws, title):
    """
    Applies professional styling to an openpyxl worksheet.
    """
    # Fonts
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Segoe UI", size=10)
    title_font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
    
    # Fills
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    accent_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    
    # Borders
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Insert Title
    ws.insert_rows(1, 2)
    ws["A1"] = title
    ws["A1"].font = title_font
    
    # Get actual columns & rows
    max_col = ws.max_column
    max_row = ws.max_row
    
    # Style Table Headers (now row 3 after insertion)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Style Data Cells
    for row in range(4, max_row + 1):
        # Zebra striping
        is_even = (row % 2 == 0)
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = cell_font
            cell.border = thin_border
            
            if is_even:
                cell.fill = accent_fill
                
            # Formatting based on value types
            val = cell.value
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            elif isinstance(val, str) and (val.startswith("202") or "-" in val or "/" in val):
                # Date format
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Calculate maximum text length
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def generate_excel_reports(metadata: dict, transactions: list[dict], df_monthly_abb: pd.DataFrame, abb_summary: dict, assessment: dict):
    """
    Generates the 4 required professional Excel reports inside the output/ folder.
    """
    output_dir = r"C:\Users\ThinkPad\Downloads\excel_file_management\ABB_Analyzer\output"
    os.makedirs(output_dir, exist_ok=True)
    
    log_info("Generating Excel reports...")
    
    try:
        # 1. Cleaned Transactions Report
        df_txs = pd.DataFrame(transactions)
        # Drop columns not needed in standard output if they exist
        cols_to_keep = [c for c in ["Date", "Particulars", "Debit", "Credit", "Balance"] if c in df_txs.columns]
        df_txs = df_txs[cols_to_keep]
        
        txs_path = os.path.join(output_dir, "cleaned_transactions.xlsx")
        with pd.ExcelWriter(txs_path, engine="openpyxl") as writer:
            df_txs.to_excel(writer, sheet_name="Transactions", index=False)
            ws = writer.sheets["Transactions"]
            format_excel_sheet(ws, f"Cleaned Transactions - {metadata.get('customer_name', 'N/A')} ({metadata.get('bank_name', 'N/A')})")
            
        # 2. Monthly ABB Report
        monthly_path = os.path.join(output_dir, "monthly_abb_report.xlsx")
        df_monthly_out = df_monthly_abb.copy()
        # Drop YearMonth helper column
        if "YearMonth" in df_monthly_out.columns:
            df_monthly_out = df_monthly_out.drop(columns=["YearMonth"])
            
        with pd.ExcelWriter(monthly_path, engine="openpyxl") as writer:
            df_monthly_out.to_excel(writer, sheet_name="Monthly ABB Details", index=False)
            ws = writer.sheets["Monthly ABB Details"]
            format_excel_sheet(ws, f"Monthly ABB Statement - Account No: {metadata.get('account_number', 'N/A')}")
            
        # 3. Overall ABB Report
        abb_path = os.path.join(output_dir, "abb_report.xlsx")
        df_abb_summary = pd.DataFrame([
            {"Metric": "Latest Month ABB (1 Month)", "Value": abb_summary["1M"]},
            {"Metric": "Average 3 Months ABB", "Value": abb_summary["3M"]},
            {"Metric": "Average 6 Months ABB", "Value": abb_summary["6M"]}
        ])
        with pd.ExcelWriter(abb_path, engine="openpyxl") as writer:
            df_abb_summary.to_excel(writer, sheet_name="ABB Summary", index=False)
            ws = writer.sheets["ABB Summary"]
            format_excel_sheet(ws, "ABB Summary Assessment")
            
        # 4. Loan Assessment Report
        assessment_path = os.path.join(output_dir, "loan_assessment.xlsx")
        # Build assessment dataframe
        df_assess = pd.DataFrame([
            {"Category": "Customer Profile", "Parameter": "Customer Name", "Details": metadata.get("customer_name", "N/A")},
            {"Category": "Customer Profile", "Parameter": "Account Number", "Details": metadata.get("account_number", "N/A")},
            {"Category": "Customer Profile", "Parameter": "Bank Name", "Details": metadata.get("bank_name", "N/A")},
            {"Category": "Statement Info", "Parameter": "Days Covered", "Details": assessment.get("days_covered", "N/A")},
            {"Category": "Balances Analysis", "Parameter": "Highest Balance", "Details": assessment.get("highest_balance", 0.0)},
            {"Category": "Balances Analysis", "Parameter": "Lowest Balance", "Details": assessment.get("lowest_balance", 0.0)},
            {"Category": "Balances Analysis", "Parameter": "Latest Balance", "Details": assessment.get("latest_balance", 0.0)},
            {"Category": "Transactions", "Parameter": "Total Credit Value", "Details": assessment.get("total_credits", 0.0)},
            {"Category": "Transactions", "Parameter": "Total Debit Value", "Details": assessment.get("total_debits", 0.0)},
            {"Category": "Transactions", "Parameter": "Credit Count", "Details": assessment.get("credit_count", 0)},
            {"Category": "Transactions", "Parameter": "Debit Count", "Details": assessment.get("debit_count", 0)},
            {"Category": "Risk Analysis", "Parameter": "ABB Status", "Details": assessment.get("abb_status", "N/A")},
            {"Category": "Risk Analysis", "Parameter": "Balance Stability", "Details": assessment.get("balance_stability", "N/A")},
            {"Category": "Risk Analysis", "Parameter": "Liquidity Strength", "Details": assessment.get("liquidity_strength", "N/A")},
            {"Category": "Risk Analysis", "Parameter": "Overall Credit Profile", "Details": assessment.get("overall_profile", "N/A")},
            {"Category": "Risk Analysis", "Parameter": "Credit Analyst Feedback", "Details": assessment.get("explanation", "N/A")}
        ])
        with pd.ExcelWriter(assessment_path, engine="openpyxl") as writer:
            df_assess.to_excel(writer, sheet_name="Credit Assessment", index=False)
            ws = writer.sheets["Credit Assessment"]
            format_excel_sheet(ws, "Credit Risk & Loan Suitability Assessment")
            
        log_info("All Excel reports generated successfully.")
    except Exception as e:
        log_error(f"Error generating Excel reports: {e}")
        raise e
